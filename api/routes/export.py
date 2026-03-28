"""Export routes - Excel/CSV download for financial data and documents."""
import io
import logging
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from database import get_db
from models.financial import FinancialStatement, FinancialLineItem
from models.user import User
from auth import get_current_user

logger = logging.getLogger(__name__)
router = APIRouter()


@router.get("/financial-statements")
def export_financial_statements(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Export all financial statements as Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    headers = ["Entity", "FY", "Type", "Revenue", "Expenditures", "Surplus/Deficit", "Fund Balance", "Debt"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    stmts = db.query(FinancialStatement).order_by(
        FinancialStatement.entity_type, FinancialStatement.fiscal_year
    ).all()

    for row, s in enumerate(stmts, 2):
        ws.cell(row=row, column=1, value=s.entity_name or s.entity_type).border = thin_border
        ws.cell(row=row, column=2, value=s.fiscal_year).border = thin_border
        ws.cell(row=row, column=3, value=s.statement_type).border = thin_border
        for col, val in enumerate([s.total_revenue, s.total_expenditures, s.surplus_deficit, s.fund_balance, s.total_debt], 4):
            cell = ws.cell(row=row, column=col, value=val)
            cell.number_format = '#,##0'
            cell.border = thin_border

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 25)

    # Line items sheet per entity
    for entity in ["town", "school"]:
        entity_stmts = [s for s in stmts if s.entity_type == entity]
        if not entity_stmts:
            continue

        ws2 = wb.create_sheet(title=f"{entity.title()} Detail")
        ws2.cell(row=1, column=1, value="Section").font = header_font
        ws2.cell(row=1, column=1).fill = header_fill
        ws2.cell(row=1, column=2, value="Line Item").font = header_font
        ws2.cell(row=1, column=2).fill = header_fill

        # Year columns
        years = sorted(set(s.fiscal_year for s in entity_stmts))
        for col, year in enumerate(years, 3):
            cell = ws2.cell(row=1, column=col, value=f"FY {year}")
            cell.font = header_font
            cell.fill = header_fill

        # Collect all unique line items across years
        all_items = {}
        for s in entity_stmts:
            items = db.query(FinancialLineItem).filter(
                FinancialLineItem.statement_id == s.id
            ).order_by(FinancialLineItem.line_order).all()
            for item in items:
                key = (item.section or "", item.line_name)
                if key not in all_items:
                    all_items[key] = {}
                all_items[key][s.fiscal_year] = item.amount

        row = 2
        current_section = ""
        for (section, name), year_data in sorted(all_items.items()):
            if section != current_section:
                current_section = section
                ws2.cell(row=row, column=1, value=section).font = Font(bold=True)
                row += 1

            ws2.cell(row=row, column=2, value=name)
            for col, year in enumerate(years, 3):
                val = year_data.get(year)
                if val is not None:
                    cell = ws2.cell(row=row, column=col, value=val)
                    cell.number_format = '#,##0'
            row += 1

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=atlantic_highlands_financials.xlsx"},
    )
